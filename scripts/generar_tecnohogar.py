from __future__ import annotations

import csv
import os
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
COURSE_ROOT = ROOT.parents[1] / "udemm_sistemas" / "seguridad_informatica"
DEFAULT_TP1 = COURSE_ROOT / "Entregas" / "TP1" / "TP1_Inventario_TecnoHogar_v2.xlsx"
TP1_WORKBOOK = Path(os.environ.get("TP1_TECNOHOGAR_XLSX", DEFAULT_TP1))
OUTPUT_DIR = ROOT / "datos" / "empresas" / "tecnohogar"
CATALOG_PATH = ROOT / "datos" / "estandares" / "iso27002_2022" / "catalogo_controles.csv"
DATE = "2026-06-04"


MATURITY = {
    0: ("0 - Inexistente", 0.00, "Sin acciones"),
    1: ("1 - Inicial", 0.05, "Esfuerzo personal"),
    2: ("2 - Gestionado", 0.15, "Buenas practicas"),
    3: ("3 - Definido", 0.60, "Procedimientos"),
    4: ("4 - Cuantitativo", 0.85, "Indicadores"),
    5: ("5 - Optimizado", 1.00, "Mejora continua"),
}

LEVELS = {
    "5.1": 3,
    "5.2": 3,
    "5.3": 2,
    "5.4": 3,
    "5.5": 2,
    "5.6": 2,
    "5.7": 1,
    "5.8": 2,
    "5.9": 4,
    "5.10": 3,
    "5.11": 3,
    "5.12": 4,
    "5.13": 3,
    "5.14": 3,
    "5.15": 3,
    "5.16": 3,
    "5.17": 3,
    "5.18": 2,
    "5.19": 2,
    "5.20": 3,
    "5.21": 2,
    "5.22": 2,
    "5.23": 2,
    "5.24": 1,
    "5.25": 2,
    "5.26": 2,
    "5.27": 1,
    "5.28": 1,
    "5.29": 2,
    "5.30": 3,
    "5.31": 4,
    "5.32": 3,
    "5.33": 4,
    "5.34": 4,
    "5.35": 2,
    "5.36": 2,
    "5.37": 3,
    "6.1": 2,
    "6.2": 3,
    "6.3": 2,
    "6.4": 3,
    "6.5": 2,
    "6.6": 3,
    "6.7": 2,
    "6.8": 3,
    "7.1": 4,
    "7.2": 3,
    "7.3": 3,
    "7.4": 2,
    "7.5": 2,
    "7.6": 2,
    "7.7": 2,
    "7.8": 3,
    "7.9": 2,
    "7.10": 2,
    "7.11": 3,
    "7.12": 3,
    "7.13": 3,
    "7.14": 2,
    "8.1": 3,
    "8.2": 1,
    "8.3": 3,
    "8.4": 2,
    "8.5": 3,
    "8.6": 3,
    "8.7": 3,
    "8.8": 1,
    "8.9": 3,
    "8.10": 3,
    "8.11": 1,
    "8.12": 1,
    "8.13": 3,
    "8.14": 3,
    "8.15": 2,
    "8.16": 1,
    "8.17": 3,
    "8.18": 1,
    "8.19": 3,
    "8.20": 3,
    "8.21": 3,
    "8.22": 3,
    "8.23": 2,
    "8.24": 3,
    "8.25": 2,
    "8.26": 2,
    "8.27": 2,
    "8.28": 1,
    "8.29": 1,
    "8.30": 2,
    "8.31": 3,
    "8.32": 3,
    "8.33": 2,
    "8.34": 3,
}

INTERVIEWEES = [
    {
        "entrevista_id": "ET-01",
        "nombre": "Juan Castillo",
        "area": "Gerencia General",
        "empresa": "TecnoHogar S.A.",
        "cargo": "Propietario de procesos comerciales y operativos",
        "fecha": DATE,
        "fuente": "TP1 roles y areas; validacion de procesos PN-01, PN-03, PN-04 y PN-05",
    },
    {
        "entrevista_id": "ET-02",
        "nombre": "Laura Mendez",
        "area": "Gerencia de Finanzas",
        "empresa": "TecnoHogar S.A.",
        "cargo": "Propietaria de pagos, proveedores e informacion financiera",
        "fecha": DATE,
        "fuente": "TP1 roles y areas; activos DI-02, DI-07, DI-09, AP-04 y AP-08",
    },
    {
        "entrevista_id": "ET-03",
        "nombre": "Susana Ortiz",
        "area": "Gerencia de RRHH",
        "empresa": "TecnoHogar S.A.",
        "cargo": "Propietaria de RRHH y liquidacion",
        "fecha": DATE,
        "fuente": "TP1 roles y areas; proceso PN-06 y activos DI-08a..DI-08d",
    },
    {
        "entrevista_id": "ET-04",
        "nombre": "Paula Ferreira",
        "area": "Gestion de Tecnologia",
        "empresa": "TecnoHogar S.A.",
        "cargo": "Propietaria de aplicaciones, servidores y bases de datos",
        "fecha": DATE,
        "fuente": "TP1 roles y areas; aplicaciones AP-01..AP-10 y servidores SV-01..SV-12",
    },
    {
        "entrevista_id": "ET-05",
        "nombre": "Diego Sosa",
        "area": "Gerencia General / Tecnologia",
        "empresa": "TecnoHogar S.A.",
        "cargo": "Custodio tecnico transversal",
        "fecha": DATE,
        "fuente": "TP1 roles y areas; operacion de e-commerce, pagos, WMS, TMS y soporte",
    },
    {
        "entrevista_id": "ET-06",
        "nombre": "Martin Rojas",
        "area": "Gerencia de RRHH / Tecnologia",
        "empresa": "TecnoHogar S.A.",
        "cargo": "Custodio de BD, backups, SIEM y activos RRHH",
        "fecha": DATE,
        "fuente": "TP1 roles y areas; BD, backups, SIEM y legajos digitales",
    },
]

PROJECTS = [
    {
        "proyecto_id": "PR-01",
        "titulo": "Gobierno SGSI, politicas y RACI",
        "plazo": "Corto",
        "esfuerzo_jornadas": 58,
        "aporte_seguridad": 0.18,
        "descripcion": "Formalizar politica de seguridad, comite, roles, revision periodica y tablero de seguimiento.",
        "dependencias": "",
        "tipo_seguridad": "Organizativa",
    },
    {
        "proyecto_id": "PR-02",
        "titulo": "Gestion de activos, clasificacion y tratamiento",
        "plazo": "Corto",
        "esfuerzo_jornadas": 46,
        "aporte_seguridad": 0.14,
        "descripcion": "Mantener el inventario TP1, extender etiquetado, reglas de uso y tratamiento de informacion critica.",
        "dependencias": "PR-01",
        "tipo_seguridad": "Organizativa",
    },
    {
        "proyecto_id": "PR-03",
        "titulo": "IAM, MFA y segregacion de funciones",
        "plazo": "Corto",
        "esfuerzo_jornadas": 72,
        "aporte_seguridad": 0.20,
        "descripcion": "Ordenar identidades, accesos privilegiados, MFA, revisiones periodicas y matriz de funciones incompatibles.",
        "dependencias": "PR-01",
        "tipo_seguridad": "Logica",
    },
    {
        "proyecto_id": "PR-04",
        "titulo": "Proveedores, cloud, privacidad y PCI-DSS",
        "plazo": "Medio",
        "esfuerzo_jornadas": 64,
        "aporte_seguridad": 0.16,
        "descripcion": "Definir anexos de seguridad, evaluacion de proveedores, gobierno cloud y obligaciones legales sobre PII y pagos.",
        "dependencias": "PR-01,PR-02",
        "tipo_seguridad": "Legal",
    },
    {
        "proyecto_id": "PR-05",
        "titulo": "Incidentes, SIEM, logs y evidencia digital",
        "plazo": "Corto",
        "esfuerzo_jornadas": 54,
        "aporte_seguridad": 0.19,
        "descripcion": "Formalizar respuesta a incidentes, clasificacion de eventos, preservacion de evidencia y monitoreo centralizado.",
        "dependencias": "PR-01",
        "tipo_seguridad": "Logica",
    },
    {
        "proyecto_id": "PR-06",
        "titulo": "Continuidad, backups y recuperacion",
        "plazo": "Medio",
        "esfuerzo_jornadas": 48,
        "aporte_seguridad": 0.13,
        "descripcion": "Definir RTO/RPO, pruebas de restauracion, continuidad TIC y criterios de seguridad durante interrupciones.",
        "dependencias": "PR-01,PR-05",
        "tipo_seguridad": "Logica",
    },
    {
        "proyecto_id": "PR-07",
        "titulo": "Seguridad fisica de oficinas y centros de distribucion",
        "plazo": "Medio",
        "esfuerzo_jornadas": 42,
        "aporte_seguridad": 0.11,
        "descripcion": "Unificar controles de acceso fisico, monitoreo, areas seguras, medios y baja segura de equipos.",
        "dependencias": "PR-01",
        "tipo_seguridad": "Fisica",
    },
    {
        "proyecto_id": "PR-08",
        "titulo": "Hardening, malware, redes y vulnerabilidades",
        "plazo": "Corto",
        "esfuerzo_jornadas": 68,
        "aporte_seguridad": 0.21,
        "descripcion": "Implantar hardening, gestion de configuracion, EDR, escaneo recurrente, segmentacion y web filtering.",
        "dependencias": "PR-03",
        "tipo_seguridad": "Logica",
    },
    {
        "proyecto_id": "PR-09",
        "titulo": "SDLC seguro, WAF, APIs y pruebas",
        "plazo": "Medio",
        "esfuerzo_jornadas": 76,
        "aporte_seguridad": 0.22,
        "descripcion": "Integrar requisitos de seguridad, codificacion segura, pruebas, separacion de ambientes y proteccion del e-commerce.",
        "dependencias": "PR-08",
        "tipo_seguridad": "Logica",
    },
    {
        "proyecto_id": "PR-10",
        "titulo": "Concientizacion y ciclo de vida RRHH",
        "plazo": "Corto",
        "esfuerzo_jornadas": 36,
        "aporte_seguridad": 0.12,
        "descripcion": "Fortalecer induccion, capacitacion, teletrabajo, confidencialidad, desvinculacion y reporte de eventos.",
        "dependencias": "PR-01",
        "tipo_seguridad": "Organizativa",
    },
    {
        "proyecto_id": "PR-11",
        "titulo": "Auditoria, metricas y mejora continua",
        "plazo": "Largo",
        "esfuerzo_jornadas": 40,
        "aporte_seguridad": 0.10,
        "descripcion": "Programar revisiones independientes, pruebas de auditoria tecnica y seguimiento de indicadores del tablero.",
        "dependencias": "PR-01,PR-05",
        "tipo_seguridad": "Organizativa",
    },
]

PROJECT_CONTROLS = {
    "PR-01": [
        "5.1",
        "5.2",
        "5.3",
        "5.4",
        "5.5",
        "5.6",
        "5.7",
        "5.8",
        "5.35",
        "5.36",
        "5.37",
    ],
    "PR-02": ["5.9", "5.10", "5.11", "5.12", "5.13", "5.14", "5.33", "5.34", "8.10", "8.11", "8.12"],
    "PR-03": ["5.15", "5.16", "5.17", "5.18", "8.2", "8.3", "8.4", "8.5", "8.18"],
    "PR-04": ["5.19", "5.20", "5.21", "5.22", "5.23", "5.31", "5.32", "5.34", "8.30"],
    "PR-05": ["5.24", "5.25", "5.26", "5.27", "5.28", "6.8", "8.15", "8.16", "8.17", "8.34"],
    "PR-06": ["5.29", "5.30", "8.6", "8.13", "8.14"],
    "PR-07": [
        "7.1",
        "7.2",
        "7.3",
        "7.4",
        "7.5",
        "7.6",
        "7.7",
        "7.8",
        "7.9",
        "7.10",
        "7.11",
        "7.12",
        "7.13",
        "7.14",
    ],
    "PR-08": ["8.1", "8.7", "8.8", "8.9", "8.19", "8.20", "8.21", "8.22", "8.23", "8.24"],
    "PR-09": ["8.25", "8.26", "8.27", "8.28", "8.29", "8.31", "8.32", "8.33"],
    "PR-10": ["6.1", "6.2", "6.3", "6.4", "6.5", "6.6", "6.7", "6.8"],
    "PR-11": ["5.27", "5.28", "5.35", "5.36", "8.15", "8.16", "8.34"],
}

CONTROL_PROJECT = {
    control_id: project_id
    for project_id, controls in PROJECT_CONTROLS.items()
    for control_id in controls
    if project_id != "PR-11"
}

EVIDENCE_OVERRIDES = {
    "5.9": "Inventario TP1 con 41 activos, propietarios, custodios, dependencias y capas PN/DI/AP/SV.",
    "5.12": "Clasificacion CID realizada sobre los 13 activos de informacion con justificacion y marco aplicable.",
    "5.24": "Los incidentes se atienden por canales operativos informales; no hay playbook aprobado ni matriz de escalamiento.",
    "5.28": "No se evidencio procedimiento de cadena de custodia ni preservacion forense para logs o evidencia digital.",
    "5.34": "Los activos DI-01, DI-02 y DI-08 registran datos personales, pago y RRHH bajo Ley 25.326, PCI-DSS y LCT.",
    "6.3": "La concientizacion existe como induccion inicial, pero no hay programa anual ni medicion de efectividad.",
    "7.1": "TP1 declara operacion web 24/7 y 3 centros de distribucion; las areas criticas estan identificadas.",
    "7.4": "Hay controles fisicos basicos en oficinas y centros, pero no monitoreo continuo homologado entre sedes.",
    "8.2": "Cuentas privilegiadas administradas por custodios historicos sin revision formal recurrente.",
    "8.8": "No hay calendario sistematico de escaneos, clasificacion de vulnerabilidades ni SLA de remediacion.",
    "8.13": "Servidor de Backup SV-09 identificado; backups configurados, sin evidencia completa de pruebas periodicas.",
    "8.15": "Servidor de Logs/SIEM SV-10 identificado; la centralizacion y retencion todavia son parciales.",
    "8.16": "Monitoreo reactivo de e-commerce y backoffice sin alertas de seguridad consistentes.",
    "8.25": "El e-commerce AP-01 es desarrollo propio, pero seguridad no esta incorporada como gate formal del ciclo.",
    "8.28": "No hay guia corporativa de codificacion segura ni checklist OWASP para cambios en AP-01.",
    "8.29": "Las pruebas de seguridad se realizan como pentest puntual, no como practica recurrente por release.",
}

OBJECTIVES = {
    "PR-01": "Alinear seguridad con direccion y habilitar gobierno medible del SGSI.",
    "PR-02": "Proteger datos criticos de clientes, pagos, stock, RRHH y registros.",
    "PR-03": "Reducir fraude, abuso de privilegios y accesos indebidos a sistemas criticos.",
    "PR-04": "Cumplir obligaciones sobre proveedores, servicios cloud, privacidad y pagos.",
    "PR-05": "Detectar, clasificar y responder eventos con trazabilidad y evidencia.",
    "PR-06": "Sostener venta online, pagos y backoffice ante fallas o interrupciones.",
    "PR-07": "Proteger oficinas, centros de distribucion, equipamiento y soportes fisicos.",
    "PR-08": "Reducir exposicion tecnica en endpoints, redes, configuraciones y malware.",
    "PR-09": "Bajar riesgo de vulnerabilidades en e-commerce, APIs e integraciones.",
    "PR-10": "Mejorar conducta segura del personal durante ingreso, operacion y egreso.",
    "PR-11": "Convertir hallazgos en seguimiento, auditoria y mejora continua.",
}


def strip_accents(value: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch))


def sorted_control_key(control_id: str) -> tuple[int, int]:
    chapter, number = control_id.split(".")
    return int(chapter), int(number)


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def load_sheet_records(workbook_path: Path, sheet: str, header_row: int) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = [cell.value for cell in ws[header_row]]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = {str(headers[index]): value for index, value in enumerate(row) if headers[index]}
        if record and record.get(str(headers[0])) not in (None, "Totales"):
            records.append(record)
    return records


def parse_refs(value: Any) -> list[str]:
    text = str(value or "").replace("—", "").replace("-", "-")
    refs: list[str] = []
    current_prefix = ""
    for token in re.split(r"[,;/]", text):
        token = token.strip()
        full = re.search(r"\b(PN|DI|AP|SV)-(\d+[a-z]?)\b", token, flags=re.IGNORECASE)
        if full:
            prefix = full.group(1).upper()
            number = full.group(2)
            current_prefix = prefix
            refs.append(f"{prefix}-{number}")
            continue
        short = re.fullmatch(r"(\d+[a-z]?)", token, flags=re.IGNORECASE)
        if short and current_prefix:
            refs.append(f"{current_prefix}-{short.group(1)}")
    return refs


def classify_asset_type(detail: str) -> str:
    normalized = strip_accents(detail).lower()
    if "proceso" in normalized:
        return "Proceso"
    if "informacion" in normalized:
        return "Informacion"
    if "aplicacion" in normalized:
        return "Aplicacion"
    return "Servidor"


def build_assets() -> list[dict[str, Any]]:
    if not TP1_WORKBOOK.exists():
        raise FileNotFoundError(
            f"No se encontro {TP1_WORKBOOK}. Definir TP1_TECNOHOGAR_XLSX con la ruta al Excel del TP1."
        )

    inventory = load_sheet_records(TP1_WORKBOOK, "Inventario", 3)
    cia_records = load_sheet_records(TP1_WORKBOOK, "Clasificación CID", 4)

    cia = {
        str(row["ID"]): {
            "confidencialidad": int(row["C"]),
            "integridad": int(row["I"]),
            "disponibilidad": int(row["D"]),
            "criticidad_cid": int(row["Máx (CID)"]),
            "justificacion_cid": row["Justificación"],
            "marco_aplicable": row["Marco Aplicable"],
        }
        for row in cia_records
        if row.get("ID")
    }

    raw_assets = []
    for row in inventory:
        asset_id = str(row["ID"])
        if not re.fullmatch(r"(PN|DI|AP|SV)-\d+[a-z]?", asset_id, flags=re.IGNORECASE):
            continue
        raw_assets.append(
            {
                "activo_id": asset_id,
                "tipo": classify_asset_type(str(row["Detalle (Tipología)"])),
                "nombre": row["Nombre del Activo"],
                "capa": row["Capa"],
                "nivel_evo": row["Nivel Evo"],
                "detalle": row["Detalle (Tipología)"],
                "descripcion": row["Descripción"],
                "propietario": row["Propietario"],
                "custodio": row["Custodio"],
                "area": row["Área"],
                "padre_principal": "" if row["Padre Principal"] == "—" else row["Padre Principal"],
                "origen": row["Origen"],
            }
        )

    by_id = {asset["activo_id"]: asset for asset in raw_assets}
    parent_refs = {asset["activo_id"]: parse_refs(asset["padre_principal"]) for asset in raw_assets}
    child_refs: dict[str, list[str]] = {}
    for child, refs in parent_refs.items():
        for parent in refs:
            child_refs.setdefault(parent, []).append(child)

    def criticality(asset_id: str, seen: set[str] | None = None) -> int:
        if asset_id in cia:
            return cia[asset_id]["criticidad_cid"]
        seen = seen or set()
        if asset_id in seen:
            return 3
        seen.add(asset_id)
        if asset_id.startswith("PN-"):
            children = child_refs.get(asset_id, [])
            values = [criticality(child, seen.copy()) for child in children if child.startswith("DI-")]
            return max(values) if values else 4
        refs = [ref for ref in parent_refs.get(asset_id, []) if ref in by_id]
        values = [criticality(ref, seen.copy()) for ref in refs]
        return max(values) if values else 4

    def process_name(asset_id: str, seen: set[str] | None = None) -> str:
        if asset_id.startswith("PN-"):
            return str(by_id[asset_id]["nombre"])
        seen = seen or set()
        if asset_id in seen:
            return ""
        seen.add(asset_id)
        for ref in parent_refs.get(asset_id, []):
            if ref in by_id:
                name = process_name(ref, seen.copy())
                if name:
                    return name
        return ""

    assets: list[dict[str, Any]] = []
    for asset in raw_assets:
        asset_id = asset["activo_id"]
        cid = cia.get(asset_id, {})
        assets.append(
            {
                **asset,
                "criticidad_cid": criticality(asset_id),
                "proceso": process_name(asset_id) or asset["nombre"],
                "confidencialidad": cid.get("confidencialidad", ""),
                "integridad": cid.get("integridad", ""),
                "disponibilidad": cid.get("disponibilidad", ""),
                "justificacion_cid": cid.get("justificacion_cid", ""),
                "marco_aplicable": cid.get("marco_aplicable", ""),
            }
        )
    return assets


def interviewee_for(control_id: str) -> str:
    chapter = int(control_id.split(".")[0])
    number = int(control_id.split(".")[1])
    if chapter == 6:
        return "Susana Ortiz"
    if chapter == 7:
        return "Martin Rojas" if number in {10, 11, 12, 13, 14} else "Juan Castillo"
    if chapter == 8:
        if number in {2, 4, 5, 8, 9, 13, 14, 15, 16, 17, 18, 20, 21, 22, 24}:
            return "Martin Rojas"
        return "Paula Ferreira"
    if control_id in {"5.19", "5.20", "5.21", "5.22", "5.23", "5.31", "5.32", "5.34"}:
        return "Laura Mendez"
    if control_id in {"5.9", "5.10", "5.11", "5.12", "5.13", "5.14", "5.33"}:
        return "Juan Castillo"
    return "Diego Sosa"


def evidence_for(control_id: str, control_name: str, level: int, project_id: str) -> tuple[str, str, str]:
    if control_id in EVIDENCE_OVERRIDES:
        evidence = EVIDENCE_OVERRIDES[control_id]
    elif level >= 4:
        evidence = f"{control_name}: existen evidencias documentadas en TP1 y seguimiento con indicadores basicos."
    elif level == 3:
        evidence = f"{control_name}: practica definida y documentada parcialmente para los procesos criticos de e-commerce."
    elif level == 2:
        evidence = f"{control_name}: buenas practicas operativas aplicadas, sin normalizacion corporativa completa."
    elif level == 1:
        evidence = f"{control_name}: tratamiento reactivo dependiente del esfuerzo de custodios tecnicos."
    else:
        evidence = f"{control_name}: no se relevaron evidencias de implementacion."

    project_title = next(project["titulo"] for project in PROJECTS if project["proyecto_id"] == project_id)
    if level >= 4:
        finding = "Control fuerte para el contexto actual; sostener evidencia y periodicidad de revision."
        observation = f"Mantener como base del tablero y usarlo para acelerar {project_title}."
    elif level == 3:
        finding = "Control definido, con oportunidad de medicion y cobertura transversal."
        observation = f"Agregar indicadores, responsables y revision periodica dentro de {project_title}."
    elif level == 2:
        finding = "Control gestionado de forma parcial; falta formalizacion, evidencia o alcance completo."
        observation = f"Priorizar normalizacion y evidencias dentro de {project_title}."
    elif level == 1:
        finding = "Control inicial y reactivo; alta dependencia de personas y bajo respaldo documental."
        observation = f"Tratar como brecha prioritaria dentro de {project_title}."
    else:
        finding = "Control inexistente para el alcance relevado."
        observation = f"Incluir definicion e implementacion desde cero dentro de {project_title}."
    return evidence, finding, observation


def build_diagnosis(catalog: list[dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for control in sorted(catalog, key=lambda row: sorted_control_key(row["control_id"])):
        control_id = control["control_id"]
        level = LEVELS[control_id]
        label, value, aspect = MATURITY[level]
        project_id = CONTROL_PROJECT.get(control_id, "PR-11")
        evidence, finding, observation = evidence_for(control_id, control["control_nombre"], level, project_id)
        rows.append(
            {
                "control_id": control_id,
                "nivel_madurez": level,
                "valor": value,
                "madurez_desc": label,
                "aspecto_clave": aspect,
                "credibilidad": "Alta",
                "cumplimiento": 1.0,
                "evidencia": evidence,
                "entrevistado": interviewee_for(control_id),
                "fecha": DATE,
                "hallazgo": finding,
                "observaciones": observation,
                "fuente_tp1": "TP1_Inventario_TecnoHogar_v2.xlsx",
                "objetivo_negocio": OBJECTIVES[project_id],
                "interpretacion": f"El grafico debe leerse contra {OBJECTIVES[project_id].lower()}",
            }
        )
    return rows


def build_projects() -> list[dict[str, Any]]:
    projects = []
    for project in PROJECTS:
        effort = float(project["esfuerzo_jornadas"])
        controls = len(PROJECT_CONTROLS[project["proyecto_id"]])
        project_counts = {"logica": 0, "fisica": 0, "organizativa": 0, "legal": 0}
        project_type = strip_accents(str(project["tipo_seguridad"])).lower()
        if project_type in project_counts:
            project_counts[project_type] = controls
        projects.append(
            {
                **project,
                **project_counts,
                "controles_esperados": controls,
                "costo": int(effort * 150000),
                "meses": round(effort / 22, 1),
            }
        )
    return projects


def build_project_control() -> list[dict[str, str]]:
    rows = []
    for project_id, controls in PROJECT_CONTROLS.items():
        for control_id in controls:
            rows.append({"proyecto_id": project_id, "control_id": control_id})
    rows.sort(key=lambda row: (row["proyecto_id"], sorted_control_key(row["control_id"])))
    return rows


def write_readme() -> None:
    (OUTPUT_DIR / "README.md").write_text(
        """# TecnoHogar

Dataset principal del TP2. Representa un diagnostico defendible de TecnoHogar S.A. sobre ISO/IEC 27002:2022, usando como base el inventario y la clasificacion del TP1.

## Fuentes

- `Entregas/TP1/TP1_Inventario_TecnoHogar_v2.xlsx`: caso TecnoHogar, 6 procesos, 41 activos, roles, dependencias y clasificacion CID.
- `datos/estandares/iso27002_2022/catalogo_controles.csv`: catalogo de 93 controles trabajado desde el ejemplo de la catedra.
- Material GESI-C6 de la catedra: criterio de tablero, KPI/KGI, madurez CMMI y plan de acciones.
- Ejemplo G1: referencia de estructura de diagnostico, madurez, proyectos y trazabilidad.

## Archivos

- `activos.csv`: 41 activos del TP1 con tipo, proceso, responsable, criticidad CID y, para datos/informacion, C/I/D y marco aplicable.
- `entrevistas.csv`: responsables usados como fuente de evidencia.
- `diagnostico.csv`: 93 controles ISO evaluados con nivel CMMI, valor normalizado, evidencia, hallazgo, observacion, fuente y objetivo de negocio.
- `proyectos.csv`: cartera de 11 iniciativas priorizables.
- `proyecto_control.csv`: relacion entre controles y proyectos.

## Criterio de numeros

La madurez global queda deliberadamente por debajo del ejemplo bancario: TecnoHogar tiene inventario, clasificacion y algunos controles operativos, pero todavia no tiene un SGSI medido y optimizado. La escala usada es CMMI 0..5:

- 0: inexistente.
- 1: inicial y reactivo.
- 2: gestionado parcialmente.
- 3: definido/documentado.
- 4: medido con indicadores.
- 5: optimizado.

## Guia para explicar graficos

- Ejecutivo: resume madurez, brecha, control mas critico, capacidad mas debil y quick wins.
- Mapa ISO: muestra cobertura de los 4 capitulos y distribucion CMMI de los 93 controles.
- Perfil: agrupa controles por funciones, dominios, CID, tipo y capacidad operacional.
- Brechas: prioriza controles por brecha ponderada; sirve para justificar el plan.
- Plan: traduce brechas a proyectos con esfuerzo, plazo, tipo de seguridad y quick wins.
- Trazabilidad: muestra la cadena capitulo -> control -> proyecto -> evidencia.

## Regeneracion

```bash
make generate-tecnohogar
```

Si el Excel TP1 esta en otra ubicacion:

```bash
TP1_TECNOHOGAR_XLSX=/ruta/TP1_Inventario_TecnoHogar_v2.xlsx make generate-tecnohogar
```
""",
        encoding="utf-8",
    )


def main() -> int:
    catalog = read_csv_dicts(CATALOG_PATH)
    assets = build_assets()
    diagnosis = build_diagnosis(catalog)
    projects = build_projects()
    project_control = build_project_control()

    write_csv(
        OUTPUT_DIR / "activos.csv",
        assets,
        [
            "activo_id",
            "tipo",
            "nombre",
            "criticidad_cid",
            "propietario",
            "proceso",
            "capa",
            "nivel_evo",
            "detalle",
            "descripcion",
            "custodio",
            "area",
            "padre_principal",
            "origen",
            "confidencialidad",
            "integridad",
            "disponibilidad",
            "justificacion_cid",
            "marco_aplicable",
        ],
    )
    write_csv(
        OUTPUT_DIR / "entrevistas.csv",
        INTERVIEWEES,
        ["entrevista_id", "nombre", "area", "empresa", "cargo", "fecha", "fuente"],
    )
    write_csv(
        OUTPUT_DIR / "diagnostico.csv",
        diagnosis,
        [
            "control_id",
            "nivel_madurez",
            "valor",
            "madurez_desc",
            "aspecto_clave",
            "credibilidad",
            "cumplimiento",
            "evidencia",
            "entrevistado",
            "fecha",
            "hallazgo",
            "observaciones",
            "fuente_tp1",
            "objetivo_negocio",
            "interpretacion",
        ],
    )
    write_csv(
        OUTPUT_DIR / "proyectos.csv",
        projects,
        [
            "proyecto_id",
            "titulo",
            "plazo",
            "esfuerzo_jornadas",
            "aporte_seguridad",
            "descripcion",
            "dependencias",
            "tipo_seguridad",
            "controles_esperados",
            "logica",
            "fisica",
            "organizativa",
            "legal",
            "costo",
            "meses",
        ],
    )
    write_csv(OUTPUT_DIR / "proyecto_control.csv", project_control, ["proyecto_id", "control_id"])
    write_readme()

    print(f"TecnoHogar generado en {OUTPUT_DIR}")
    print(f"- activos: {len(assets)}")
    print(f"- entrevistas: {len(INTERVIEWEES)}")
    print(f"- diagnostico: {len(diagnosis)} controles")
    print(f"- proyectos: {len(projects)}")
    print(f"- vinculos control-proyecto: {len(project_control)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
