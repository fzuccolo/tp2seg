from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
MATERIA = ROOT.parent.parent / "udemm_sistemas" / "seguridad_informatica"
EJEMPLO_XLSX = MATERIA / "Recursos" / "TP1 2025 G1" / "REF G1 - Metricas_ISO_27002_2022_Grupo2_FINAL.xlsx"
PROYECTOS_XLSX = MATERIA / "Recursos" / "TP2 Material" / "Relacion Controles Proyectos ISO27K.xlsx"

STANDARD_DIR = ROOT / "datos" / "estandares" / "iso27002_2022"
COMPANY_DIR = ROOT / "datos" / "empresas" / "ejemplo"


CAPACITY_COLUMNS = {
    19: "cap_gobernanza",
    20: "cap_gestion_activos",
    21: "cap_proteccion_informacion",
    22: "cap_seguridad_rrhh",
    23: "cap_seguridad_fisica",
    24: "cap_seguridad_redes_sistemas",
    25: "cap_seguridad_aplicaciones",
    26: "cap_configuraciones_seguras",
    27: "cap_gestion_accesos_identidades",
    28: "cap_gestion_amenazas_vulnerabilidades",
    29: "cap_continuidad",
    30: "cap_seguridad_proveedores",
    31: "cap_cumplimiento_legalidad",
    32: "cap_gestion_eventos_si",
    33: "cap_garantizar_si",
}

DOMAIN_COLUMNS = {
    34: "dom_gobernanza_ecosistema",
    35: "dom_proteccion",
    36: "dom_defensa",
    37: "dom_resiliencia",
}


def value(cell: Any) -> Any:
    if cell is None:
        return ""
    if isinstance(cell, str):
        return " ".join(cell.replace("\r", "\n").split())
    return cell


def text(cell: Any) -> str:
    return str(value(cell)).strip()


def number(cell: Any, default: float = 0.0) -> float:
    cell = value(cell)
    if cell == "":
        return default
    try:
        return float(cell)
    except (TypeError, ValueError):
        return default


def integer(cell: Any, default: int = 0) -> int:
    return int(round(number(cell, float(default))))


def control_id(metric: str) -> str:
    match = re.match(r"\s*(\d+\.\d+)", metric)
    if not match:
        raise ValueError(f"No se pudo obtener control_id desde {metric!r}")
    return match.group(1)


def control_name(metric: str) -> str:
    return re.sub(r"^\s*\d+\.\d+\s*-\s*", "", metric).strip()


def maturity_level(maturity: str) -> int:
    match = re.match(r"\s*(\d+)", maturity or "")
    if not match:
        return 0
    return int(match.group(1))


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def load_workbooks() -> tuple[Any, Any]:
    if not EJEMPLO_XLSX.exists():
        raise FileNotFoundError(EJEMPLO_XLSX)
    if not PROYECTOS_XLSX.exists():
        raise FileNotFoundError(PROYECTOS_XLSX)
    return (
        openpyxl.load_workbook(EJEMPLO_XLSX, data_only=True),
        openpyxl.load_workbook(PROYECTOS_XLSX, data_only=True),
    )


def extract_control_rows(wb: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    cuestionario = wb["Cuestionario"]
    anexo = wb["Ctrl_Anexo"]

    catalog: list[dict[str, Any]] = []
    diagnostic: list[dict[str, Any]] = []
    links: list[dict[str, Any]] = []

    for row in range(5, 98):
        metric = text(anexo.cell(row, 2).value)
        if not metric:
            continue
        cid = control_id(metric)
        chapter = text(anexo.cell(row, 6).value)
        project_title = text(anexo.cell(row, 51).value)
        plazo = text(anexo.cell(row, 52).value)

        item = {
            "control_id": cid,
            "capitulo": chapter,
            "control_nombre": control_name(metric),
            "control_descripcion": text(anexo.cell(row, 3).value),
            "pregunta": text(anexo.cell(row, 4).value),
            "proposito": text(anexo.cell(row, 5).value),
            "tipo_preventivo": integer(anexo.cell(row, 7).value),
            "tipo_detectivo": integer(anexo.cell(row, 8).value),
            "tipo_correctivo": integer(anexo.cell(row, 9).value),
            "prop_confidencialidad": integer(anexo.cell(row, 10).value),
            "prop_integridad": integer(anexo.cell(row, 11).value),
            "prop_disponibilidad": integer(anexo.cell(row, 12).value),
            "func_identificacion": integer(anexo.cell(row, 13).value),
            "func_proteccion": integer(anexo.cell(row, 14).value),
            "func_deteccion": integer(anexo.cell(row, 15).value),
            "func_respuesta": integer(anexo.cell(row, 16).value),
            "func_recuperacion": integer(anexo.cell(row, 17).value),
            "aplica": integer(anexo.cell(row, 49).value, 1),
            "peso": round(number(anexo.cell(row, 46).value, 1.0), 6),
            "porcentaje_absoluto_capitulo": round(number(anexo.cell(row, 43).value), 6),
            "tipo_seguridad": text(anexo.cell(row, 50).value),
            "proyecto_sugerido": project_title,
            "plazo_sugerido": plazo,
            "mapeo_iso_27002_2013": text(anexo.cell(row, 53).value),
        }
        for col, name in CAPACITY_COLUMNS.items():
            item[name] = integer(anexo.cell(row, col).value)
        for col, name in DOMAIN_COLUMNS.items():
            item[name] = integer(anexo.cell(row, col).value)
        catalog.append(item)

        diagnostic.append(
            {
                "control_id": cid,
                "nivel_madurez": maturity_level(text(cuestionario.cell(row, 8).value)),
                "valor": number(cuestionario.cell(row, 9).value),
                "madurez_desc": text(cuestionario.cell(row, 8).value),
                "aspecto_clave": text(cuestionario.cell(row, 10).value),
                "credibilidad": text(cuestionario.cell(row, 11).value),
                "cumplimiento": number(cuestionario.cell(row, 12).value),
                "evidencia": text(cuestionario.cell(row, 16).value),
                "entrevistado": text(cuestionario.cell(row, 13).value),
                "fecha": text(cuestionario.cell(row, 14).value).split(" ")[0],
                "hora": text(cuestionario.cell(row, 15).value),
                "hallazgo": text(cuestionario.cell(row, 16).value),
                "observacion": text(cuestionario.cell(row, 17).value),
                "comentarios": text(cuestionario.cell(row, 18).value),
                "observaciones": text(cuestionario.cell(row, 17).value) or text(cuestionario.cell(row, 18).value),
            }
        )

        if project_title:
            links.append(
                {
                    "proyecto_titulo": project_title,
                    "control_id": cid,
                    "plazo": plazo,
                    "tipo_seguridad": text(anexo.cell(row, 50).value),
                    "brecha_control": round(number(anexo.cell(row, 48).value), 6),
                }
            )

    return catalog, diagnostic, links


def extract_interviews(wb: Any) -> list[dict[str, Any]]:
    ws = wb["Entrevistas"]
    rows: list[dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        interview_id = text(ws.cell(row, 2).value)
        if not interview_id:
            continue
        rows.append(
            {
                "entrevista_id": interview_id,
                "nombre": text(ws.cell(row, 3).value),
                "area": text(ws.cell(row, 4).value),
                "empresa": text(ws.cell(row, 5).value),
                "cargo": text(ws.cell(row, 6).value),
            }
        )
    return rows


def extract_projects(project_wb: Any, links: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ws = project_wb["Proyectos"]
    headers = [text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    source: list[dict[str, Any]] = []
    by_title: dict[str, dict[str, Any]] = {}

    for row in range(2, ws.max_row + 1):
        raw = {headers[col - 1]: ws.cell(row, col).value for col in range(1, ws.max_column + 1) if headers[col - 1]}
        code = text(raw.get("Código"))
        title = text(raw.get("Título"))
        if not code or not title:
            continue
        item = {
            "proyecto_id": code,
            "titulo": title,
            "plazo": text(raw.get("Plazo_Std")),
            "esfuerzo_jornadas": number(raw.get("Jornadas")),
            "aporte_seguridad": round(number(raw.get("Peso_Abs")) / 100, 6),
            "descripcion": text(raw.get("Descripción")),
            "dependencias": text(raw.get("Padre")),
            "tipo_seguridad": text(raw.get("Tipo_Seg")),
            "controles_esperados": integer(raw.get("Ctrl_Esp")),
            "logica": integer(raw.get("Lógica")),
            "fisica": integer(raw.get("Física")),
            "organizativa": integer(raw.get("Organizativa")),
            "legal": integer(raw.get("Legal")),
            "costo": number(raw.get("Costo")),
            "meses": number(raw.get("Meses")),
        }
        source.append(item)
        by_title[title] = item

    used_titles = {link["proyecto_titulo"] for link in links}
    for title in sorted(used_titles - set(by_title)):
        code = f"PEX{len(source) + 1:02d}"
        item = {
            "proyecto_id": code,
            "titulo": title,
            "plazo": next((link["plazo"] for link in links if link["proyecto_titulo"] == title), ""),
            "esfuerzo_jornadas": 0,
            "aporte_seguridad": 0,
            "descripcion": "Proyecto identificado en el anexo del ejemplo G1.",
            "dependencias": "",
            "tipo_seguridad": next((link["tipo_seguridad"] for link in links if link["proyecto_titulo"] == title), ""),
            "controles_esperados": 0,
            "logica": 0,
            "fisica": 0,
            "organizativa": 0,
            "legal": 0,
            "costo": 0,
            "meses": 0,
        }
        source.append(item)
        by_title[title] = item

    project_links = []
    seen: set[tuple[str, str]] = set()
    for link in links:
        project = by_title[link["proyecto_titulo"]]
        key = (project["proyecto_id"], link["control_id"])
        if key in seen:
            continue
        seen.add(key)
        project_links.append(
            {
                "proyecto_id": project["proyecto_id"],
                "control_id": link["control_id"],
                "plazo": link["plazo"],
                "tipo_seguridad": link["tipo_seguridad"],
                "brecha_control": link["brecha_control"],
            }
        )

    return source, project_links


def write_metadata() -> None:
    (COMPANY_DIR / "empresa.yml").write_text(
        "\n".join(
            [
                "id: ejemplo",
                "nombre: Banco de Buenos Aires S.A.",
                "descripcion: Caso ejemplo G1 usado como referencia de tablero ISO 27002:2022.",
                "industria: Servicios financieros",
                "estandar: iso27002_2022",
                "fuente_principal: REF G1 - Metricas_ISO_27002_2022_Grupo2_FINAL.xlsx",
                "",
            ]
        ),
        encoding="utf-8",
    )

    write_csv(
        COMPANY_DIR / "activos.csv",
        [
            {
                "activo_id": "ORG-01",
                "tipo": "Organizacion",
                "nombre": "Banco de Buenos Aires S.A.",
                "criticidad_cid": 5,
                "propietario": "CISO",
                "proceso": "Sistema de Gestion de Seguridad de la Informacion",
            },
            {
                "activo_id": "PR-01",
                "tipo": "Proceso",
                "nombre": "Gestion estrategica de seguridad",
                "criticidad_cid": 5,
                "propietario": "CISO",
                "proceso": "Plan estrategico de seguridad",
            },
        ],
        ["activo_id", "tipo", "nombre", "criticidad_cid", "propietario", "proceso"],
    )


def main() -> int:
    wb, project_wb = load_workbooks()
    catalog, diagnostic, raw_links = extract_control_rows(wb)
    interviews = extract_interviews(wb)
    projects, project_links = extract_projects(project_wb, raw_links)

    write_csv(
        STANDARD_DIR / "catalogo_controles.csv",
        catalog,
        [
            "control_id",
            "capitulo",
            "control_nombre",
            "control_descripcion",
            "pregunta",
            "proposito",
            "tipo_preventivo",
            "tipo_detectivo",
            "tipo_correctivo",
            "prop_confidencialidad",
            "prop_integridad",
            "prop_disponibilidad",
            "func_identificacion",
            "func_proteccion",
            "func_deteccion",
            "func_respuesta",
            "func_recuperacion",
            *CAPACITY_COLUMNS.values(),
            *DOMAIN_COLUMNS.values(),
            "aplica",
            "peso",
            "porcentaje_absoluto_capitulo",
            "tipo_seguridad",
            "proyecto_sugerido",
            "plazo_sugerido",
            "mapeo_iso_27002_2013",
        ],
    )

    write_csv(
        COMPANY_DIR / "diagnostico.csv",
        diagnostic,
        [
            "control_id",
            "nivel_madurez",
            "valor",
            "madurez_desc",
            "aspecto_clave",
            "credibilidad",
            "cumplimiento",
            "evidencia",
            "entrevistado",
            "fecha",
            "hora",
            "hallazgo",
            "observacion",
            "comentarios",
            "observaciones",
        ],
    )

    write_csv(
        COMPANY_DIR / "entrevistas.csv",
        interviews,
        ["entrevista_id", "nombre", "area", "empresa", "cargo"],
    )

    write_csv(
        COMPANY_DIR / "proyectos.csv",
        projects,
        [
            "proyecto_id",
            "titulo",
            "plazo",
            "esfuerzo_jornadas",
            "aporte_seguridad",
            "descripcion",
            "dependencias",
            "tipo_seguridad",
            "controles_esperados",
            "logica",
            "fisica",
            "organizativa",
            "legal",
            "costo",
            "meses",
        ],
    )

    write_csv(
        COMPANY_DIR / "proyecto_control.csv",
        project_links,
        ["proyecto_id", "control_id", "plazo", "tipo_seguridad", "brecha_control"],
    )

    write_metadata()

    print(f"Controles importados: {len(catalog)}")
    print(f"Diagnosticos importados: {len(diagnostic)}")
    print(f"Entrevistas importadas: {len(interviews)}")
    print(f"Proyectos importados: {len(projects)}")
    print(f"Vinculos proyecto-control importados: {len(project_links)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
